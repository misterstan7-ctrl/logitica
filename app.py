from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file

from database import (
    criar_banco,
    conectar,
    salvar_entrega,
    dados_dashboard,
    dados_relatorio,
    salvar_ids_pacotes,
    listar_ids_por_data,
    buscar_id_pacote,
    total_ids_data
)

from io import BytesIO

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter


app = Flask(__name__)
app.secret_key = "troque-essa-chave-depois"

USUARIO_PADRAO = "admin"
SENHA_PADRAO = "1234"


# ==========================================
# INICIAR BANCO
# ==========================================
@app.before_request
def iniciar():
    criar_banco()


# ==========================================
# LOGIN
# ==========================================
def login_obrigatorio():
    return session.get("logado") == True


# ==========================================
# LOGIN
# ==========================================
@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "").strip()

        if usuario == USUARIO_PADRAO and senha == SENHA_PADRAO:

            session["logado"] = True
            session["usuario"] = usuario

            return redirect(url_for("dashboard"))

        flash("Usuário ou senha inválidos.", "erro")

    return render_template("login.html")


# ==========================================
# DASHBOARD
# ==========================================
@app.route("/dashboard")
def dashboard():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    dados = dados_dashboard()

    return render_template(
        "dashboard.html",
        dados=dados
    )


# ==========================================
# ADICIONAR PACOTES
# ==========================================
@app.route("/adicionar", methods=["GET", "POST"])
def adicionar():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    if request.method == "POST":

        data = request.form.get("data")
        recebidos = request.form.get("recebidos")
        entregues = request.form.get("entregues")

        try:

            recebidos = int(recebidos)
            entregues = int(entregues)

            if recebidos < 0 or entregues < 0:

                flash(
                    "Os valores não podem ser negativos.",
                    "erro"
                )

                return redirect(url_for("adicionar"))

            if entregues > recebidos:

                flash(
                    "Entregues não pode ser maior que recebidos.",
                    "erro"
                )

                return redirect(url_for("adicionar"))

            salvar_entrega(
                data,
                recebidos,
                entregues
            )

            flash(
                "Pacotes salvos com sucesso!",
                "sucesso"
            )

            return redirect(url_for("dashboard"))

        except:

            flash(
                "Preencha os campos corretamente.",
                "erro"
            )

            return redirect(url_for("adicionar"))

    return render_template("adicionar.html")

@app.route("/buscar-id", methods=["GET", "POST"])
def buscar_id():
    if not login_obrigatorio():
        return redirect(url_for("login"))

    resultados = []
    codigo = ""

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()

        if not codigo:
            flash("Digite um ID para buscar.", "erro")
            return redirect(url_for("buscar_id"))

        resultados = buscar_id_pacote(codigo)

        if not resultados:
            flash("Nenhum pacote encontrado com esse ID.", "erro")

    return render_template(
        "buscar_id.html",
        resultados=resultados,
        codigo=codigo
    )
# ==========================================
# ADICIONAR IDS
# ==========================================
@app.route("/adicionar-ids", methods=["GET", "POST"])
def adicionar_ids():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    ids_salvos = []
    data_selecionada = ""
    total = 0

    if request.method == "POST":

        data = request.form.get("data")
        ids_texto = request.form.get("ids", "")

        lista_ids = [
            i.strip()
            for i in ids_texto.splitlines()
            if i.strip()
        ]

        # ==========================================
        # BUSCA PACOTES DO DIA
        # ==========================================
        conn = conectar()
        cur = conn.cursor()

        cur.execute("""
            SELECT recebidos
            FROM entregas
            WHERE data = ?
        """, (data,))

        resultado = cur.fetchone()

        conn.close()

        if not resultado:

            flash(
                "Nenhum registro de pacotes encontrado nessa data.",
                "erro"
            )

            return redirect(url_for("adicionar_ids"))

        recebidos = int(resultado["recebidos"])

        # ==========================================
        # VALIDAR QUANTIDADE IDS
        # ==========================================
        if len(lista_ids) > recebidos:

            flash(
                f"Existem MAIS IDs ({len(lista_ids)}) do que pacotes recebidos ({recebidos}).",
                "erro"
            )

            return redirect(url_for("adicionar_ids"))

        if len(lista_ids) < recebidos:

            faltando = recebidos - len(lista_ids)

            flash(
                f"Faltam {faltando} IDs para completar os pacotes do dia.",
                "erro"
            )

            return redirect(url_for("adicionar_ids"))

        # ==========================================
        # SALVAR IDS
        # ==========================================
        erros = salvar_ids_pacotes(
            data,
            lista_ids
        )

        if erros:

            flash(
                f"{len(erros)} ID(s) duplicado(s) encontrados.",
                "erro"
            )

        else:

            flash(
                "IDs adicionados com sucesso!",
                "sucesso"
            )

        ids_salvos = listar_ids_por_data(data)

        data_selecionada = data
        total = total_ids_data(data)

    return render_template(
        "adicionar_ids.html",
        ids_salvos=ids_salvos,
        data_selecionada=data_selecionada,
        total=total
    )


# ==========================================
# RELATORIOS
# ==========================================
@app.route("/relatorios")
def relatorios():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    periodo = request.args.get("periodo", "mes")
    inicio = request.args.get("inicio")
    fim = request.args.get("fim")

    dados = dados_relatorio(
        periodo,
        inicio,
        fim
    )

    return render_template(
        "relatorios.html",
        dados=dados
    )


# ==========================================
# EXPORTAR PDF
# ==========================================
@app.route("/exportar/pdf")
def exportar_pdf():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    periodo = request.args.get("periodo", "mes")
    inicio = request.args.get("inicio")
    fim = request.args.get("fim")

    dados = dados_relatorio(
        periodo,
        inicio,
        fim
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4
    )

    styles = getSampleStyleSheet()

    elementos = []

    titulo = Paragraph(
        "<b>Relatório de Entregas</b>",
        styles["Title"]
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 20))

    tabela_dados = [
        [
            "Data",
            "Recebidos",
            "Entregues",
            "Taxa",
            "Valor"
        ]
    ]

    for linha in dados["linhas"]:

        tabela_dados.append([
            linha["data"],
            str(linha["recebidos"]),
            str(linha["entregues"]),
            f"{linha['taxa']:.2f}%",
            f"R$ {linha['valor']:.2f}"
        ])

    tabela = Table(tabela_dados)

    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elementos.append(tabela)

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio.pdf",
        mimetype="application/pdf"
    )


# ==========================================
# EXPORTAR EXCEL
# ==========================================
@app.route("/exportar/excel")
def exportar_excel():

    if not login_obrigatorio():
        return redirect(url_for("login"))

    periodo = request.args.get("periodo", "mes")
    inicio = request.args.get("inicio")
    fim = request.args.get("fim")

    dados = dados_relatorio(
        periodo,
        inicio,
        fim
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Relatório"

    headers = [
        "Data",
        "Recebidos",
        "Entregues",
        "Taxa",
        "Valor"
    ]

    ws.append(headers)

    for linha in dados["linhas"]:

        ws.append([
            linha["data"],
            linha["recebidos"],
            linha["entregues"],
            linha["taxa"],
            linha["valor"]
        ])

    azul = "1D4ED8"

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=azul
        )

    thin = Side(
        border_style="thin",
        color="CBD5E1"
    )

    for row in ws.iter_rows():

        for cell in row:

            cell.border = Border(
                top=thin,
                left=thin,
                right=thin,
                bottom=thin
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

    for col in range(1, 6):

        ws.column_dimensions[
            get_column_letter(col)
        ].width = 20

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================================
# SAIR
# ==========================================
@app.route("/sair")
def sair():

    session.clear()

    return redirect(url_for("login"))


# ==========================================
# FILTROS
# ==========================================
@app.template_filter("moeda")
def moeda(valor):

    return (
        f"R$ {valor:,.2f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


@app.template_filter("porcentagem")
def porcentagem(valor):

    return (
        f"{valor:.2f}%"
        .replace(".", ",")
    )


# ==========================================
# INICIAR
# ==========================================
import os

if __name__ == "__main__":
    criar_banco()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
